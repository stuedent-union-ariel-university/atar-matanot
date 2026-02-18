#!/usr/bin/env python3
"""
Script: Upload users from an XLSX file to a Monday board.

Input file format (first sheet):
 - Column A: user id (required)
 - Column B: user name (optional but recommended)
 - Header row will be skipped by default; pass --no-header if there is no header.

Env defaults (override via CLI flags):
 - MONDAY_API_KEY
 - MONDAY_USER_BOARD_ID                       (--board)
 - MONDAY_USER_BOARD_USER_ID_COLUMN_ID       (--user-id-col)
 - MONDAY_USER_BOARD_USER_NAME_COLUMN_ID     (--user-name-col)

Examples:
  python scripts/upload_users_from_xlsx.py --file ./users.xlsx
  python scripts/upload_users_from_xlsx.py --file ./users.xlsx --board 12345 --user-id-col text --user-name-col text1
  python scripts/upload_users_from_xlsx.py --file ./users.xlsx --dry  # preview only
"""

from __future__ import annotations

import argparse
import json
import os
import random
import sys
import time
from io import BytesIO
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook


MONDAY_API_URL = "https://api.monday.com/v2"


def to_bool(v: Optional[str], default: bool = False) -> bool:
    if v is None:
        return default
    s = str(v).strip().lower()
    if s in {"true", "1", "yes", "y"}:
        return True
    if s in {"false", "0", "no", "n"}:
        return False
    return default


def require_env(name: str, friendly: Optional[str] = None) -> str:
    val = os.getenv(name)
    if not val:
        raise RuntimeError(f"Missing env: {friendly or name}")
    return val


@dataclass
class RetryConfig:
    retries: int = 7
    min_delay_ms: int = 250
    max_delay_ms: int = 5000
    jitter_ms: int = 250


def jitter_sleep(base_ms: int, jitter_ms: int) -> None:
    j = random.randint(0, max(0, jitter_ms))
    time.sleep((base_ms + j) / 1000.0)


def monday_request_with_retry(
    query: str,
    variables: Dict,
    api_key: str,
    retry_cfg: RetryConfig,
    timeout_s: int = 30,
) -> Dict:
    """
    Minimal retry wrapper for Monday GraphQL.
    Retries on network errors and on API errors (including rate limiting),
    using exponential backoff with jitter, similar spirit to your TS helper.
    """
    headers = {
        "Authorization": api_key,
        "Content-Type": "application/json",
    }

    last_err: Optional[Exception] = None

    for attempt in range(retry_cfg.retries + 1):
        try:
            resp = requests.post(
                MONDAY_API_URL,
                headers=headers,
                json={"query": query, "variables": variables},
                timeout=timeout_s,
            )

            # Non-200 could happen (rate limit / gateway etc.)
            if resp.status_code != 200:
                raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:500]}")

            payload = resp.json()

            # Monday GraphQL errors come here
            if "errors" in payload and payload["errors"]:
                # Sometimes rate limit errors appear in errors[].message
                msg = payload["errors"][0].get("message", "Unknown Monday error")
                raise RuntimeError(f"Monday GraphQL error: {msg}")

            return payload

        except Exception as e:
            last_err = e
            if attempt >= retry_cfg.retries:
                break

            # exponential backoff capped
            # attempt=0 -> min_delay, attempt=1 -> 2x, etc.
            backoff = retry_cfg.min_delay_ms * (2**attempt)
            backoff = min(backoff, retry_cfg.max_delay_ms)
            jitter_sleep(backoff, retry_cfg.jitter_ms)

    raise RuntimeError(f"Request failed after retries: {last_err}") from last_err


def validate_file(file_path: str) -> str:
    abs_path = os.path.abspath(file_path)
    if not os.path.exists(abs_path):
        raise FileNotFoundError(f"File not found: {abs_path}")
    lower = abs_path.lower()
    if not (
        lower.endswith(".xlsx")
        or lower.endswith(".xlsm")
        or lower.endswith(".xltx")
        or lower.endswith(".xltm")
    ):
        print("Warning: file does not look like an Excel workbook.", file=sys.stderr)
    return abs_path


def decrypt_workbook_if_needed(
    file_path: str,
    password: Optional[str],
    decrypt_to: Optional[str],
) -> str:
    if not password:
        return file_path

    try:
        import msoffcrypto
    except Exception as exc:
        raise RuntimeError(
            "Password-protected XLSX requires msoffcrypto-tool. "
            "Install with: pip install msoffcrypto-tool"
        ) from exc

    if decrypt_to:
        out_path = os.path.abspath(decrypt_to)
    else:
        root, ext = os.path.splitext(file_path)
        out_path = f"{root}.decrypted{ext}"

    with open(file_path, "rb") as in_f:
        office_file = msoffcrypto.OfficeFile(in_f)
        office_file.load_key(password=password)
        decrypted = BytesIO()
        office_file.decrypt(decrypted)

    with open(out_path, "wb") as out_f:
        out_f.write(decrypted.getvalue())

    return out_path


def read_users_from_workbook(
    file_path: str,
    skip_header: bool = True,
) -> Iterable[Tuple[str, Optional[str]]]:
    """
    Yields (user_id, user_name_or_none) from first worksheet, columns A/B.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    row_iter = ws.iter_rows(values_only=True)
    required_department = 'בה"ס לרפואה'

    for idx, row in enumerate(row_iter):
        if skip_header and idx == 0:
            continue

        user_id = str(row[0] if row and len(row) > 0 else "").strip()
        user_name = str(row[1] if row and len(row) > 1 else "").strip()
        department = str(row[3] if row and len(row) > 3 else "").strip()

        if not user_id:
            continue

        if department != required_department:
            continue

        yield user_id, (user_name if user_name else None)


CREATE_ITEM_MUTATION = """
mutation($boardId: ID!, $itemName: String!, $columnValues: JSON!, $groupId: String) {
  create_item(board_id: $boardId, item_name: $itemName, column_values: $columnValues, group_id: $groupId) { id }
}
"""


def create_user_item(
    *,
    api_key: str,
    board_id: str,
    user_id_col: str,
    user_id: str,
    user_name_col: Optional[str],
    user_name: Optional[str],
    group_id: Optional[str],
    retry_cfg: RetryConfig,
) -> Dict:
    values: Dict[str, str] = {user_id_col: str(user_id)}
    if user_name_col and user_name:
        values[user_name_col] = str(user_name)

    variables = {
        "boardId": str(board_id),
        "itemName": (user_name.strip() if user_name else "") or str(user_id),
        "columnValues": json.dumps(values),
        "groupId": group_id,
    }

    return monday_request_with_retry(
        CREATE_ITEM_MUTATION, variables, api_key=api_key, retry_cfg=retry_cfg
    )


def chunked(lst: List, n: int) -> Iterable[List]:
    for i in range(0, len(lst), n):
        yield lst[i : i + n]


def main() -> int:
    load_dotenv()

    p = argparse.ArgumentParser()
    p.add_argument("--file", required=True, help="path/to/users.xlsx")

    p.add_argument(
        "--board", help="Monday board id (overrides env MONDAY_USER_BOARD_ID)"
    )
    p.add_argument(
        "--user-id-col",
        dest="user_id_col",
        help="Column id for user id (overrides env MONDAY_USER_BOARD_USER_ID_COLUMN_ID)",
    )
    p.add_argument(
        "--user-name-col",
        dest="user_name_col",
        help="Column id for user name (overrides env MONDAY_USER_BOARD_USER_NAME_COLUMN_ID)",
    )
    p.add_argument("--group", help="Optional group id")
    p.add_argument("--password", help="Workbook password for encrypted XLSX")
    p.add_argument(
        "--decrypt-to",
        help=(
            "Path for decrypted workbook. Defaults to <file>.decrypted.xlsx when --password is used."
        ),
    )

    p.add_argument(
        "--dry",
        nargs="?",
        const="true",
        default=None,
        help="Dry run (default true if flag present). Use --dry=false to disable.",
    )
    p.add_argument(
        "--no-header",
        nargs="?",
        const="true",
        default=None,
        help="Treat first row as data (default true if flag present).",
    )

    p.add_argument(
        "--limit", type=int, default=0, help="Max rows to process (0 means no limit)"
    )
    p.add_argument(
        "--batch", type=int, default=10, help="Batch size (sequential batches)"
    )
    p.add_argument(
        "--delayMs", type=int, default=100, help="Delay between batches (ms)"
    )

    p.add_argument("--retries", type=int, default=7)
    p.add_argument("--minDelayMs", type=int, default=250)
    p.add_argument("--maxDelayMs", type=int, default=5000)
    p.add_argument("--jitterMs", type=int, default=250)

    args = p.parse_args()

    abs_file = validate_file(args.file)
    abs_file = decrypt_workbook_if_needed(abs_file, args.password, args.decrypt_to)

    api_key = require_env("MONDAY_API_KEY")

    board_id = args.board or os.getenv("MONDAY_USER_BOARD_ID", "")
    user_id_col = args.user_id_col or os.getenv(
        "MONDAY_USER_BOARD_USER_ID_COLUMN_ID", ""
    )
    user_name_col = args.user_name_col or os.getenv(
        "MONDAY_USER_BOARD_USER_NAME_COLUMN_ID", ""
    )
    group_id = args.group

    if not board_id:
        raise RuntimeError(
            "Board id is required. Provide --board or MONDAY_USER_BOARD_ID."
        )
    if not user_id_col:
        raise RuntimeError(
            "User id column id is required. Provide --user-id-col or MONDAY_USER_BOARD_USER_ID_COLUMN_ID."
        )

    dry = to_bool(args.dry, default=False)
    no_header = to_bool(args.no_header, default=False)

    retry_cfg = RetryConfig(
        retries=args.retries,
        min_delay_ms=args.minDelayMs,
        max_delay_ms=args.maxDelayMs,
        jitter_ms=args.jitterMs,
    )

    # Read + dedupe by user id (keep first occurrence)
    seen = set()
    users: List[Tuple[str, Optional[str]]] = []
    for uid, name in read_users_from_workbook(
        abs_file,
        skip_header=(not no_header),
    ):
        if uid in seen:
            continue
        seen.add(uid)
        users.append((uid, name))

    total = len(users)
    if args.limit and args.limit > 0:
        total = min(total, args.limit)

    if total == 0:
        print("No user rows found to process.")
        return 0

    print(f"Preparing to upload {total} user(s) to board {board_id}")
    cols_line = f"Using columns -> id: {user_id_col}"
    if user_name_col:
        cols_line += f", name: {user_name_col}"
    if group_id:
        cols_line += f", group: {group_id}"
    print(cols_line)

    if dry:
        print("Dry-run. First 10 rows:")
        for uid, name in users[: min(10, total)]:
            print(f" - id: {uid}\tname: {name or ''}")
        print("Pass --dry=false or omit --dry to execute.")
        return 0

    created = 0
    # sequential batches; within each batch we do sequential calls (safe/simple).
    # If you want real parallelism like Promise.allSettled, we can add ThreadPoolExecutor.
    for batch_users in chunked(users[:total], max(1, args.batch)):
        ok = 0
        fail = 0
        first_err: Optional[Exception] = None

        for uid, name in batch_users:
            try:
                create_user_item(
                    api_key=api_key,
                    board_id=str(board_id),
                    user_id_col=str(user_id_col),
                    user_id=str(uid),
                    user_name_col=(str(user_name_col) if user_name_col else None),
                    user_name=name,
                    group_id=group_id,
                    retry_cfg=retry_cfg,
                )
                ok += 1
            except Exception as e:
                fail += 1
                if first_err is None:
                    first_err = e

        created += ok
        if fail > 0 and first_err:
            print(f"\nSome creations failed (batch): {first_err}", file=sys.stderr)

        if args.delayMs:
            time.sleep(args.delayMs / 1000.0)

        sys.stdout.write(f"\rCreated: {created}/{total}")
        sys.stdout.flush()

    print("\nDone.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as err:
        print(f"Failed to upload users from xlsx: {err}", file=sys.stderr)
        raise SystemExit(1)
